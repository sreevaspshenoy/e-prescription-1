from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import certifi
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
from contextlib import asynccontextmanager
import uuid
from datetime import datetime, timezone
import jwt
import hashlib
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import BaseDocTemplate, Frame, PageTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import urllib.request

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection (use certifi CA bundle for Atlas SSL on macOS)
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url, tlsCAFile=certifi.where())
db = client[os.environ['DB_NAME']]

# JWT Secret
JWT_SECRET = os.environ.get('JWT_SECRET', 'rheumacare-secret-key-2024')
JWT_ALGORITHM = "HS256"

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Startup: init doctors. Shutdown: close Mongo client."""
    try:
        await init_doctors()
    except Exception as e:
        logger.warning("Startup init_doctors failed (app will still serve): %s", e)
    yield
    client.close()


# Create the main app
app = FastAPI(lifespan=lifespan)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Admin credentials (hardcoded for security)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "rheumacare_admin_2024"

# Hash password helper
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed: str) -> bool:
    return hash_password(password) == hashed

# Default doctors data (will be migrated to DB)
DEFAULT_DOCTORS = [
    {
        "id": "dr_prakashini",
        "name": "Dr. Prakashini M V",
        "qualifications": "MD, DM (Clinical Immunology & Rheumatology)",
        "role": "Consultant Rheumatologist",
        "kmc_no": "KMC No. 113674",
        "location": "Bangalore",
        "username": "prakashini",
        "password_hash": hash_password("prakashini123"),
        "is_active": True
    },
    {
        "id": "dr_dharmanand",
        "name": "Dr. B.G Dharmanand",
        "qualifications": "MD, DM (Rheumatology)",
        "role": "Consultant Rheumatologist",
        "kmc_no": "KMC No. 75545",
        "location": "Bangalore",
        "username": "dharmanand",
        "password_hash": hash_password("dharmanand123"),
        "is_active": True
    },
    {
        "id": "dr_ramesh",
        "name": "Dr. Ramesh Jois",
        "qualifications": "MD, MRCP, CCST (Rheum, UK)",
        "role": "Consultant Rheumatologist",
        "kmc_no": "KMC No. 33688",
        "location": "Bangalore",
        "username": "ramesh",
        "password_hash": hash_password("ramesh123"),
        "is_active": True
    }
]

# Legacy DOCTORS dict for backward compatibility
DOCTORS = {d["id"]: d for d in DEFAULT_DOCTORS}

# Clinic Info (shared across all doctors)
CLINIC_INFO = {
    "clinic_name": "Rheuma CARE",
    "slogan": "Live Pain-Free",
    "address": "1/1, GF & FF, Millers Arcade, Millers Road, Vasanth Nagar, Bangalore, Karnataka, Pin: 560052",
    "timing": "8:30 AM - 6:30 PM",
    "phone_appointments": "080-35000700",
    "phone_helpline": "080-35000701"
}

# Keep DOCTOR_INFO for backward compatibility
DOCTOR_INFO = {**DOCTORS["dr_prakashini"], **CLINIC_INFO}

# Drug list (from CSV - extracted drug names without dosages)
DRUG_LIST = [
    "WYSOLONE", "REFRESH TEARS", "ECOSPRIN", "ECOSPRIN AV", "GLYCOMET",
    "NUCOXIA", "ACTEMRA", "ALLEGRA", "TRYPTOMER", "PRODEP", "GLYCOMET GP1",
    "MEDROL", "LIMCEE", "NEXITO", "TECZINE", "NEXITO PLUS", "THYRONORM",
    "RISOFOS", "ROZAVEL", "MYORIL", "UPRISE D3", "OSTEOFOS", "PREGABA M",
    "OMNACORTIL", "TELMA", "PREGALIN", "RIFAGUT", "ALTRADAY", "BENADON",
    "MONTEK LC", "MYOSPAS", "DULANE", "BACTRIM DS", "MUVERA", "R-CINEX",
    "GABAWIN", "SOMPRAZ D", "NEUROBION FORTE", "NERVIJEN PLUS", "SOMPRAZ",
    "TOFE", "MGD3", "PANTOCID", "ULTRACET", "ME 12 OD", "LIVOGEN Z",
    "DEFCORT", "NICARDIA RETARD", "BECOSULES", "LOSAR", "NUSAM", "OMEZ",
    "PRIXAIN GEL", "SHELCAL-M", "LONAZEP", "AZORAN", "V B7 HAIR", "ACITROM",
    "SHELCAL HD", "ANXIT", "RETOZ", "PREDMET", "PILOMAX", "ULTRA MAGNESIUM",
    "GOLIMUREL", "BETRECEP", "ASSURANS", "PREGABA", "PREGABA NT",
    "AUGMENTIN DUO", "EVION LC", "OTRIVIN NASAL SPRAY", "ACTIGUT", "LIVOGEN",
    "UPADOZ", "SHELCAL", "NAPROSYN", "CILACAR", "HISONE", "CCM TAB",
    "LEVOFLOX", "PREVENAR 13 VACCINE", "ZYCOLCHIN", "GABANTIN", "ZYCEL",
    "MICROCID", "FOL-5MG", "SHELCAL XT", "GABANTIN NT", "CREMAFFIN PLUS",
    "FLOMIST NASAL SPRAY", "NALTOX", "EVION", "AB PHYLLINE N", "D VENIZ",
    "IDROFOS", "VENUSIA MAX LOTION", "DABIGO", "ZINCOVIT", "APEXFER",
    "FOLLIHAIR NEW", "PROLAGE PLUS", "AMLONG", "DOTHIP", "T BACT OINT",
    "METHORA PFS", "PNEUMOVAX 23 PFS VIAL", "SALACTIN PAINT",
    "PREVENAR 20 VACCINE", "XYKAA BD", "VERTIN", "A TO Z NS+", "ENVAS",
    "LUBRIJOINT OD", "ARISTOZYME", "SOLONEX", "SENSIVAL", "CILACAR T",
    "EMESET", "RESTYL", "OXALGIN NANOGEL", "FERISOME", "UDILIV", "OROFER XT",
    "BILASURE", "BILAST", "SUPRADYN", "TENVIR AF", "RABLET", "FOLLIHAIR",
    "DOLO", "OMALIREL", "HAEM UP GEMS", "FLEXCART", "LEVOCET", "ACTON OR",
    "PRUTIS", "EXEMPTIA", "PREDNIWIK", "FOLINAL PLUS", "PYRIGESIC", "SEACOD",
    "ONE M-D3", "NORMAL SALINE", "CINTODAC", "DENOSTEOREL", "AZEE",
    "INTACEPT", "JOGREN", "PULMONEXT", "MIMOD", "ROZEL", "METHORA",
    "INFLUVAC TETRA", "EBUXO", "SAAZ", "PEG-NT", "IGUVIC", "MMF S",
    "CONIMUNE M", "FOLITRAX", "STOPLOS A+", "HCQS", "MMF", "APRAIZE",
    "NEO-DROL", "MYCOMUNE S", "MYCOMUNE", "IGURATI", "PEG SR M", "TACROMUS",
    "TFCT-NIB", "TACROCORD", "LEFNO", "TADACT", "SAAZ DS", "PENTALOC D",
    "OSTEOCAL"
]

# Models
class DrugEntry(BaseModel):
    drug_name: str
    dosage: str
    frequency: str
    duration: str
    duration_unit: str
    comments: str = ""

class PrescriptionCreate(BaseModel):
    op_no: str
    patient_name: str
    sex: str = ""
    age: str = ""
    icd_code: str = ""
    # Individual vitals fields
    weight: str = ""
    height: str = ""
    bp: str = ""
    spo2: str = ""
    diagnosis: str
    clinical_history: str
    drugs: List[DrugEntry]
    review_after: str
    advice: str = ""
    lab_tests: str = ""
    doctor_id: str = "dr_prakashini"
    location: str = "Bangalore"

class Prescription(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    op_no: str
    patient_name: str
    sex: str = ""
    age: str = ""
    icd_code: str = ""
    # Individual vitals fields
    weight: str = ""
    height: str = ""
    bp: str = ""
    spo2: str = ""
    diagnosis: str
    clinical_history: str
    drugs: List[DrugEntry]
    review_after: str
    advice: str = ""
    lab_tests: str = ""
    doctor_id: str = "dr_prakashini"
    location: str = "Bangalore"
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

# Doctor models
class DoctorCreate(BaseModel):
    name: str
    qualifications: str = ""
    role: str = "Consultant Rheumatologist"
    kmc_no: str = ""
    location: str
    username: str
    password: str
    is_active: bool = True

class DoctorUpdate(BaseModel):
    name: Optional[str] = None
    qualifications: Optional[str] = None
    role: Optional[str] = None
    kmc_no: Optional[str] = None
    location: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None
    is_active: Optional[bool] = None

class DoctorResponse(BaseModel):
    id: str
    name: str
    qualifications: str
    role: str
    kmc_no: str
    location: str
    username: str
    is_active: bool

class LoginRequest(BaseModel):
    username: str
    password: str

class LoginResponse(BaseModel):
    token: str
    user: str
    role: str
    doctor_id: Optional[str] = None
    location: Optional[str] = None

# Auth helper - returns payload with role info
def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# Admin check helper
def require_admin(payload: dict = Depends(verify_token)):
    if payload.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return payload

# Initialize doctors in database on startup
async def init_doctors():
    """Initialize default doctors if not present in database"""
    count = await db.doctors.count_documents({})
    if count == 0:
        for doctor in DEFAULT_DOCTORS:
            await db.doctors.insert_one(doctor)
        logging.info(f"Initialized {len(DEFAULT_DOCTORS)} default doctors")

# Routes
@api_router.get("/")
async def root():
    return {"message": "RheumaCare E-Prescription API"}

@api_router.post("/auth/login", response_model=LoginResponse)
async def login(request: LoginRequest):
    # Check admin login
    if request.username == ADMIN_USERNAME and request.password == ADMIN_PASSWORD:
        token = jwt.encode(
            {
                "user": request.username, 
                "role": "admin",
                "exp": datetime.now(timezone.utc).timestamp() + 86400
            },
            JWT_SECRET,
            algorithm=JWT_ALGORITHM
        )
        return {"token": token, "user": request.username, "role": "admin"}
    
    # Check doctor login from database
    doctor = await db.doctors.find_one({"username": request.username, "is_active": True}, {"_id": 0})
    if doctor and verify_password(request.password, doctor.get("password_hash", "")):
        token = jwt.encode(
            {
                "user": request.username,
                "role": "doctor",
                "doctor_id": doctor["id"],
                "location": doctor.get("location", "Bangalore"),
                "exp": datetime.now(timezone.utc).timestamp() + 86400
            },
            JWT_SECRET,
            algorithm=JWT_ALGORITHM
        )
        return {
            "token": token, 
            "user": doctor["name"], 
            "role": "doctor",
            "doctor_id": doctor["id"],
            "location": doctor.get("location", "Bangalore")
        }
    
    raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.get("/auth/verify")
async def verify_auth(payload: dict = Depends(verify_token)):
    return {
        "valid": True, 
        "user": payload.get("user"),
        "role": payload.get("role", "doctor"),
        "doctor_id": payload.get("doctor_id"),
        "location": payload.get("location")
    }

# Doctor Management (Admin only)
@api_router.get("/admin/doctors")
async def get_all_doctors(payload: dict = Depends(require_admin)):
    """Get all doctors (admin only)"""
    doctors = await db.doctors.find({}, {"_id": 0, "password_hash": 0}).to_list(100)
    return {"doctors": doctors}

@api_router.post("/admin/doctors")
async def create_doctor(doctor: DoctorCreate, payload: dict = Depends(require_admin)):
    """Create a new doctor (admin only)"""
    # Check if username already exists
    existing = await db.doctors.find_one({"username": doctor.username})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    doctor_id = f"dr_{doctor.username.lower().replace(' ', '_')}"
    
    # Check if ID already exists
    existing_id = await db.doctors.find_one({"id": doctor_id})
    if existing_id:
        doctor_id = f"dr_{doctor.username.lower().replace(' ', '_')}_{str(uuid.uuid4())[:8]}"
    
    doctor_doc = {
        "id": doctor_id,
        "name": doctor.name,
        "qualifications": doctor.qualifications,
        "role": doctor.role,
        "kmc_no": doctor.kmc_no,
        "location": doctor.location,
        "username": doctor.username,
        "password_hash": hash_password(doctor.password),
        "is_active": doctor.is_active,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.doctors.insert_one(doctor_doc)
    
    # Return without password_hash
    del doctor_doc["password_hash"]
    if "_id" in doctor_doc:
        del doctor_doc["_id"]
    return doctor_doc

@api_router.put("/admin/doctors/{doctor_id}")
async def update_doctor(doctor_id: str, doctor: DoctorUpdate, payload: dict = Depends(require_admin)):
    """Update a doctor (admin only)"""
    existing = await db.doctors.find_one({"id": doctor_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Doctor not found")
    
    update_data = {}
    if doctor.name is not None:
        update_data["name"] = doctor.name
    if doctor.qualifications is not None:
        update_data["qualifications"] = doctor.qualifications
    if doctor.role is not None:
        update_data["role"] = doctor.role
    if doctor.kmc_no is not None:
        update_data["kmc_no"] = doctor.kmc_no
    if doctor.location is not None:
        update_data["location"] = doctor.location
    if doctor.username is not None:
        # Check if new username is taken by another doctor
        other = await db.doctors.find_one({"username": doctor.username, "id": {"$ne": doctor_id}})
        if other:
            raise HTTPException(status_code=400, detail="Username already taken")
        update_data["username"] = doctor.username
    if doctor.password is not None:
        update_data["password_hash"] = hash_password(doctor.password)
    if doctor.is_active is not None:
        update_data["is_active"] = doctor.is_active
    
    if update_data:
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.doctors.update_one({"id": doctor_id}, {"$set": update_data})
    
    updated = await db.doctors.find_one({"id": doctor_id}, {"_id": 0, "password_hash": 0})
    return updated

@api_router.delete("/admin/doctors/{doctor_id}")
async def delete_doctor(doctor_id: str, payload: dict = Depends(require_admin)):
    """Delete a doctor (admin only)"""
    result = await db.doctors.delete_one({"id": doctor_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Doctor not found")
    return {"message": "Doctor deleted successfully"}

# Public doctors endpoint (for dropdown, returns active doctors only)
@api_router.get("/doctors")
async def get_doctors(payload: dict = Depends(verify_token)):
    """Get active doctors for dropdown"""
    doctors = await db.doctors.find(
        {"is_active": True}, 
        {"_id": 0, "password_hash": 0}
    ).to_list(100)
    
    # If no doctors in DB, return from default
    if not doctors:
        doctors = [{k: v for k, v in d.items() if k != "password_hash"} for d in DEFAULT_DOCTORS]
    
    return {"doctors": doctors}

@api_router.get("/doctors/{doctor_id}")
async def get_doctor(doctor_id: str, payload: dict = Depends(verify_token)):
    doctor = await db.doctors.find_one({"id": doctor_id}, {"_id": 0, "password_hash": 0})
    if doctor:
        return {**doctor, **CLINIC_INFO}
    # Fallback to default doctors
    if doctor_id in DOCTORS:
        return {**DOCTORS[doctor_id], **CLINIC_INFO}
    raise HTTPException(status_code=404, detail="Doctor not found")

@api_router.get("/drugs")
async def get_drugs(search: Optional[str] = None, payload: dict = Depends(verify_token)):
    if search:
        filtered = [d for d in DRUG_LIST if search.upper() in d.upper()]
        return {"drugs": filtered[:20]}
    return {"drugs": DRUG_LIST[:50]}

@api_router.get("/doctor-info")
async def get_doctor_info():
    return DOCTOR_INFO

@api_router.post("/prescriptions", response_model=Prescription)
async def create_prescription(prescription: PrescriptionCreate, payload: dict = Depends(verify_token)):
    # Get doctor's location if not provided
    location = prescription.location
    if payload.get("role") == "doctor":
        location = payload.get("location", "Bangalore")
    
    prescription_obj = Prescription(
        op_no=prescription.op_no,
        patient_name=prescription.patient_name,
        sex=prescription.sex,
        age=prescription.age,
        icd_code=prescription.icd_code,
        weight=prescription.weight,
        height=prescription.height,
        bp=prescription.bp,
        spo2=prescription.spo2,
        diagnosis=prescription.diagnosis,
        clinical_history=prescription.clinical_history,
        drugs=[d.model_dump() for d in prescription.drugs],
        review_after=prescription.review_after,
        advice=prescription.advice,
        lab_tests=prescription.lab_tests,
        doctor_id=prescription.doctor_id,
        location=location
    )
    
    doc = prescription_obj.model_dump()
    await db.prescriptions.insert_one(doc)
    return prescription_obj

@api_router.get("/prescriptions", response_model=List[Prescription])
async def get_prescriptions(payload: dict = Depends(verify_token)):
    # Admin sees all, doctors see only their own
    query = {}
    if payload.get("role") == "doctor":
        query["doctor_id"] = payload.get("doctor_id")
    
    prescriptions = await db.prescriptions.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return prescriptions

@api_router.get("/prescriptions/export/excel")
async def export_prescriptions_excel(payload: dict = Depends(verify_token)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    
    # Admin sees all, doctors see only their own
    query = {}
    if payload.get("role") == "doctor":
        query["doctor_id"] = payload.get("doctor_id")
    
    prescriptions = await db.prescriptions.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Prescriptions"
    
    # Headers - added Location column
    headers = ["OP No", "Patient Name", "Sex", "Age", "ICD Code", "Weight", "Height", "BP", "SpO2", "Date", "Drug Name", "Dosage", "Frequency", "Duration", "Comments", "Advice", "Lab Tests", "Doctor", "Location"]
    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.border = thin_border
    
    # Data rows
    row_num = 2
    for prescription in prescriptions:
        drugs = prescription.get('drugs', [])
        op_no = prescription.get('op_no', '')
        patient_name = prescription.get('patient_name', '')
        sex = prescription.get('sex', '')
        age = prescription.get('age', '')
        icd_code = prescription.get('icd_code', '')
        weight = prescription.get('weight', '')
        height = prescription.get('height', '')
        bp = prescription.get('bp', '')
        spo2 = prescription.get('spo2', '')
        created_at = prescription.get('created_at', '')[:10] if prescription.get('created_at') else ''
        advice = prescription.get('advice', '')
        lab_tests = prescription.get('lab_tests', '')
        doctor_id = prescription.get('doctor_id', 'dr_prakashini')
        doctor_name = DOCTORS.get(doctor_id, {}).get('name', 'Dr. Prakashini M V')
        location = prescription.get('location', 'Bangalore')
        
        if drugs:
            for i, drug in enumerate(drugs):
                ws.cell(row=row_num, column=1, value=op_no if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=2, value=patient_name if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=3, value=sex if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=4, value=age if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=5, value=icd_code if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=6, value=weight if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=7, value=height if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=8, value=bp if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=9, value=spo2 if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=10, value=created_at if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=11, value=drug.get('drug_name', '')).border = thin_border
                ws.cell(row=row_num, column=12, value=drug.get('dosage', '')).border = thin_border
                ws.cell(row=row_num, column=13, value=drug.get('frequency', '')).border = thin_border
                ws.cell(row=row_num, column=14, value=f"{drug.get('duration', '')} {drug.get('duration_unit', '')}").border = thin_border
                ws.cell(row=row_num, column=15, value=drug.get('comments', '')).border = thin_border
                ws.cell(row=row_num, column=16, value=advice if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=17, value=lab_tests if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=18, value=doctor_name if i == 0 else '').border = thin_border
                ws.cell(row=row_num, column=19, value=location if i == 0 else '').border = thin_border
                row_num += 1
        else:
            ws.cell(row=row_num, column=1, value=op_no).border = thin_border
            ws.cell(row=row_num, column=2, value=patient_name).border = thin_border
            ws.cell(row=row_num, column=3, value=sex).border = thin_border
            ws.cell(row=row_num, column=4, value=age).border = thin_border
            ws.cell(row=row_num, column=5, value=icd_code).border = thin_border
            ws.cell(row=row_num, column=6, value=weight).border = thin_border
            ws.cell(row=row_num, column=7, value=height).border = thin_border
            ws.cell(row=row_num, column=8, value=bp).border = thin_border
            ws.cell(row=row_num, column=9, value=spo2).border = thin_border
            ws.cell(row=row_num, column=10, value=created_at).border = thin_border
            ws.cell(row=row_num, column=11, value='').border = thin_border
            ws.cell(row=row_num, column=12, value='').border = thin_border
            ws.cell(row=row_num, column=13, value='').border = thin_border
            ws.cell(row=row_num, column=14, value='').border = thin_border
            ws.cell(row=row_num, column=15, value='').border = thin_border
            ws.cell(row=row_num, column=16, value=advice).border = thin_border
            ws.cell(row=row_num, column=17, value=lab_tests).border = thin_border
            ws.cell(row=row_num, column=18, value=doctor_name).border = thin_border
            ws.cell(row=row_num, column=19, value=location).border = thin_border
            row_num += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['S'].width = 15  # Location column
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 20
    ws.column_dimensions['M'].width = 25
    ws.column_dimensions['N'].width = 25
    ws.column_dimensions['O'].width = 20
    ws.column_dimensions['L'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=prescriptions_{datetime.now().strftime('%Y%m%d')}.xlsx"
        }
    )

@api_router.get("/prescriptions/{prescription_id}", response_model=Prescription)
async def get_prescription(prescription_id: str, payload: dict = Depends(verify_token)):
    prescription = await db.prescriptions.find_one({"id": prescription_id}, {"_id": 0})
    if not prescription:
        raise HTTPException(status_code=404, detail="Prescription not found")
    return prescription

@api_router.get("/prescriptions/by-op/{op_no}")
async def get_prescriptions_by_op(op_no: str, payload: dict = Depends(verify_token)):
    """Get all prescriptions for a given OP number"""
    prescriptions = await db.prescriptions.find({"op_no": op_no}, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"prescriptions": prescriptions}

@api_router.delete("/prescriptions/{prescription_id}")
async def delete_prescription(prescription_id: str, payload: dict = Depends(verify_token)):
    """Delete a prescription by ID"""
    result = await db.prescriptions.delete_one({"id": prescription_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Prescription not found")
    return {"message": "Prescription deleted successfully"}

@api_router.put("/prescriptions/{prescription_id}", response_model=Prescription)
async def update_prescription(prescription_id: str, prescription: PrescriptionCreate, payload: dict = Depends(verify_token)):
    """Update an existing prescription"""
    # Check if prescription exists
    existing = await db.prescriptions.find_one({"id": prescription_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Prescription not found")
    
    # Update the prescription
    update_data = {
        "op_no": prescription.op_no,
        "patient_name": prescription.patient_name,
        "sex": prescription.sex,
        "age": prescription.age,
        "icd_code": prescription.icd_code,
        "weight": prescription.weight,
        "height": prescription.height,
        "bp": prescription.bp,
        "spo2": prescription.spo2,
        "diagnosis": prescription.diagnosis,
        "clinical_history": prescription.clinical_history,
        "drugs": [d.model_dump() for d in prescription.drugs],
        "review_after": prescription.review_after,
        "advice": prescription.advice,
        "lab_tests": prescription.lab_tests,
        "doctor_id": prescription.doctor_id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.prescriptions.update_one({"id": prescription_id}, {"$set": update_data})
    
    # Return updated prescription
    updated = await db.prescriptions.find_one({"id": prescription_id}, {"_id": 0})
    return updated

@api_router.get("/prescriptions/{prescription_id}/pdf")
async def generate_pdf(prescription_id: str, debug: bool = False, payload: dict = Depends(verify_token)):
    prescription = await db.prescriptions.find_one({"id": prescription_id}, {"_id": 0})
    if not prescription:
        raise HTTPException(status_code=404, detail="Prescription not found")
    
    # Get doctor info (stored for history, but not printed on PDF since using pre-printed pads)
    doctor_id = prescription.get('doctor_id', 'dr_prakashini')
    doctor_info = DOCTORS.get(doctor_id, DOCTORS['dr_prakashini'])
    
    # Generate PDF - designed for pre-printed prescription pads
    # Header and footer spaces are reserved but not printed
    buffer = BytesIO()
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    label_style = ParagraphStyle(
        'Label',
        parent=styles['Normal'],
        fontSize=10,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#333333')
    )
    
    content_style = ParagraphStyle(
        'Content',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333')
    )
    
    rx_style = ParagraphStyle(
        'Rx',
        parent=styles['Normal'],
        fontSize=14,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#6B9A9A')
    )
    
    # Page dimensions and margins for pre-printed prescription pads
    # Header: 36mm, Footer: 36mm, Padding: 4mm on each side
    page_width, page_height = A4  # 210mm x 297mm
    left_margin = 15*mm
    right_margin = 15*mm
    header_space = 36*mm  # Reserved for pre-printed header
    footer_space = 36*mm  # Reserved for pre-printed footer
    padding = 4*mm        # Padding between header/footer and content
    
    # Total margins = header/footer space + padding
    top_margin = header_space + padding      # 40mm total from top
    bottom_margin = footer_space + padding   # 40mm total from bottom
    
    # Frame dimensions - content area between margins
    frame_width = page_width - left_margin - right_margin
    frame_height = page_height - top_margin - bottom_margin  # 217mm usable height
    
    # Define page callback for debug mode - draws margin lines on every page
    def draw_debug_margins(canvas, doc):
        if debug:
            canvas.saveState()
            # Draw header zone (36mm from top)
            canvas.setStrokeColor(colors.blue)
            canvas.setLineWidth(1)
            header_line_y = page_height - header_space
            canvas.line(0, header_line_y, page_width, header_line_y)
            
            # Draw footer zone (36mm from bottom)
            footer_line_y = footer_space
            canvas.line(0, footer_line_y, page_width, footer_line_y)
            
            # Draw content boundary lines (with 4mm padding)
            canvas.setStrokeColor(colors.red)
            top_content_y = page_height - top_margin  # 36mm + 4mm = 40mm from top
            bottom_content_y = bottom_margin          # 36mm + 4mm = 40mm from bottom
            canvas.line(0, top_content_y, page_width, top_content_y)
            canvas.line(0, bottom_content_y, page_width, bottom_content_y)
            
            # Add labels
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.blue)
            canvas.drawString(5, header_line_y + 2, f"HEADER: 36mm")
            canvas.drawString(5, footer_line_y + 2, f"FOOTER: 36mm")
            canvas.setFillColor(colors.red)
            canvas.drawString(5, top_content_y - 10, f"CONTENT TOP (36mm + 4mm padding)")
            canvas.drawString(5, bottom_content_y + 12, f"CONTENT BOTTOM (36mm + 4mm padding)")
            canvas.restoreState()
    
    # Create document with BaseDocTemplate for precise margin control
    doc = BaseDocTemplate(buffer, pagesize=A4,
                         leftMargin=left_margin, rightMargin=right_margin,
                         topMargin=top_margin, bottomMargin=bottom_margin)
    
    # Create frame - positioned at (left_margin, bottom_margin) with calculated dimensions
    # Content will be top-aligned within this frame
    frame = Frame(left_margin, bottom_margin, frame_width, frame_height,
                  leftPadding=0, bottomPadding=0, rightPadding=0, topPadding=0,
                  id='main')
    
    # Page template with debug callback
    template = PageTemplate(id='main', frames=[frame], onPage=draw_debug_margins)
    doc.addPageTemplates([template])
    
    elements = []
    
    # Date and Patient Info Row - NEW LAYOUT: Patient Name on left, OP No/Sex/Age on right
    created_date = datetime.fromisoformat(prescription['created_at'].replace('Z', '+00:00'))
    date_str = created_date.strftime("%d-%m-%Y")
    
    # First row: Patient Name on left, Date and OP No on right
    sex = prescription.get('sex', '')
    age = prescription.get('age', '')
    icd_code = prescription.get('icd_code', '')
    
    info_row = Table([
        [
            Paragraph(f"<b>Patient Name:</b> {prescription['patient_name']}", content_style),
            Paragraph(f"<b>Date:</b> {date_str}  |  <b>OP No.:</b> {prescription['op_no']}", ParagraphStyle('RightAlign', parent=content_style, alignment=TA_RIGHT))
        ]
    ], colWidths=[90*mm, 90*mm])
    info_row.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(info_row)
    elements.append(Spacer(1, 3*mm))
    
    # Second row: ICD Code on left, Sex and Age on right
    if sex or age or icd_code:
        right_parts = []
        if sex:
            right_parts.append(f"<b>Sex:</b> {sex}")
        if age:
            right_parts.append(f"<b>Age:</b> {age}")
        right_text = "  |  ".join(right_parts) if right_parts else ""
        
        info_row2 = Table([
            [
                Paragraph(f"<b>ICD Code:</b> {icd_code}" if icd_code else "", content_style),
                Paragraph(right_text, ParagraphStyle('RightAlign', parent=content_style, alignment=TA_RIGHT))
            ]
        ], colWidths=[90*mm, 90*mm])
        info_row2.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ]))
        elements.append(info_row2)
    
    # Vitals row - individual fields formatted: Wt, Ht, BP, SpO2
    weight = prescription.get('weight', '')
    height = prescription.get('height', '')
    bp = prescription.get('bp', '')
    spo2 = prescription.get('spo2', '')
    
    if weight or height or bp or spo2:
        elements.append(Spacer(1, 2*mm))
        vitals_parts = []
        if weight:
            vitals_parts.append(f"<b>Wt:</b> {weight}")
        if height:
            vitals_parts.append(f"<b>Ht:</b> {height}")
        if bp:
            vitals_parts.append(f"<b>BP:</b> {bp}")
        if spo2:
            vitals_parts.append(f"<b>SpO2:</b> {spo2}")
        vitals_text = "  |  ".join(vitals_parts)
        elements.append(Paragraph(f"<b>Vitals:</b> {vitals_text}", content_style))
    
    elements.append(Spacer(1, 8*mm))
    
    # Diagnosis with proper spacing
    elements.append(Paragraph("<b>Diagnosis:</b>", label_style))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(f"    {prescription['diagnosis']}", ParagraphStyle('DiagContent', parent=content_style, leading=14)))
    elements.append(Spacer(1, 8*mm))
    
    # Clinical History with proper spacing
    if prescription.get('clinical_history'):
        elements.append(Paragraph("<b>Clinical History:</b>", label_style))
        elements.append(Spacer(1, 2*mm))
        elements.append(Paragraph(f"    {prescription['clinical_history']}", ParagraphStyle('HistContent', parent=content_style, leading=14)))
        elements.append(Spacer(1, 8*mm))
    
    # Rx with proper spacing
    elements.append(Paragraph("Rx :", rx_style))
    elements.append(Spacer(1, 5*mm))
    
    # Drug Table - increased frequency column width
    # Using Paragraphs for comments to enable text wrapping
    comment_style = ParagraphStyle(
        'Comment',
        parent=content_style,
        fontSize=9,
        leading=11,
        wordWrap='CJK'  # Enable word wrapping
    )
    
    drug_header = ['S.No', 'Drug Name', 'Dosage', 'Frequency', 'Duration', 'Comments']
    drug_data = [drug_header]
    
    for idx, drug in enumerate(prescription['drugs'], 1):
        # Wrap comments in Paragraph for proper text wrapping
        comment_text = drug.get('comments', '-') or '-'
        comment_para = Paragraph(comment_text, comment_style)
        
        drug_data.append([
            str(idx),
            drug['drug_name'],
            drug['dosage'],
            drug['frequency'],
            f"{drug['duration']} {drug['duration_unit']}",
            comment_para
        ])
    
    # Adjusted column widths - frequency column wider
    # splitByRow=1 allows table to split between rows for better page filling
    drug_table = Table(drug_data, colWidths=[10*mm, 38*mm, 22*mm, 35*mm, 25*mm, 40*mm],
                       repeatRows=1, splitByRow=1)  # Repeat header and allow row splitting
    drug_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B9A9A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (1, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 3*mm),
        ('TOPPADDING', (0, 0), (-1, 0), 3*mm),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 2.5*mm),
        ('TOPPADDING', (0, 1), (-1, -1), 2.5*mm),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Top align for wrapped text
    ]))
    elements.append(drug_table)
    elements.append(Spacer(1, 8*mm))
    
    # Review After
    if prescription.get('review_after'):
        elements.append(Paragraph(f"<b>Review After:</b> {prescription['review_after']}", content_style))
        elements.append(Spacer(1, 5*mm))
    
    # Advice section
    if prescription.get('advice'):
        elements.append(Spacer(1, 3*mm))
        elements.append(Paragraph("<b>Advice / Instructions:</b>", label_style))
        elements.append(Spacer(1, 2*mm))
        elements.append(Paragraph(f"    {prescription['advice']}", ParagraphStyle('AdviceContent', parent=content_style, leading=14)))
        elements.append(Spacer(1, 5*mm))
    
    # Lab Tests section
    if prescription.get('lab_tests'):
        elements.append(Spacer(1, 3*mm))
        elements.append(Paragraph("<b>Lab Tests Advised:</b>", label_style))
        elements.append(Spacer(1, 2*mm))
        elements.append(Paragraph(f"    {prescription['lab_tests']}", ParagraphStyle('LabContent', parent=content_style, leading=14)))
        elements.append(Spacer(1, 5*mm))
    
    # No signature printed - using pre-printed prescription pads with doctor signature
    # Doctor info is stored in database and shown in history, but not on printed PDF
    
    # Build the document - SimpleDocTemplate automatically handles 50mm margins on all pages
    doc.build(elements)
    
    buffer.seek(0)
    pdf_bytes = buffer.getvalue()
    
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=prescription_{prescription['op_no']}_{date_str}.pdf"
        }
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
